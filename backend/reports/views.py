"""
Reports API Views
Provides analytics and reporting data for admin dashboard
"""

import csv
import io
from datetime import timedelta

from accommodation.models import AccommodationRequest
from accounts.models import User
from accounts.utils import has_permission
from bookings.models import FlightBooking, HotelBooking
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from transport.models import TransportRequest
from trf.models import TravelRequest
from utils.api_response import error_response, forbidden_response, success_response
from visa.models import VisaApplication
from workflows.models import WorkflowInstance, WorkflowStepExecution


def _require_admin_reports_permission(request):
    """Shared gate for the admin reports/analytics endpoints below.
    Returns a 403 response if the user lacks generate_admin_reports, else None.

    export_data holders are also let through: ReportExportView (below) calls
    straight into these same view classes to fetch the underlying report
    data before exporting it, so an export_data-only user (e.g. Line
    Manager) needs to pass this gate too, not just the export endpoint's
    own check."""
    if (
        request.user.is_superuser
        or has_permission(request.user, "generate_admin_reports")
        or has_permission(request.user, "export_data")
    ):
        return None
    return forbidden_response(
        message="You do not have permission to view admin reports"
    )


class AdminReportsView(APIView):
    """
    Main reports endpoint for admin analytics
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get comprehensive report data
        Query params:
        - date_range: week, month, quarter, year (default: month)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        date_range = request.query_params.get("date_range", "month")

        # Calculate date range
        now = timezone.now()
        if date_range == "week":
            start_date = now - timedelta(days=7)
        elif date_range == "quarter":
            start_date = now - timedelta(days=90)
        elif date_range == "year":
            start_date = now - timedelta(days=365)
        else:  # month
            start_date = now - timedelta(days=30)

        # Get all requests within date range
        trf_requests = TravelRequest.objects.filter(created_at__gte=start_date)
        transport_requests = TransportRequest.objects.filter(created_at__gte=start_date)
        visa_requests = VisaApplication.objects.filter(created_at__gte=start_date)
        accommodation_requests = AccommodationRequest.objects.filter(
            created_at__gte=start_date
        )

        # Calculate key metrics
        total_requests = (
            trf_requests.count()
            + transport_requests.count()
            + visa_requests.count()
            + accommodation_requests.count()
        )

        # Get previous period for comparison
        previous_start = start_date - (now - start_date)
        previous_total = (
            TravelRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            ).count()
            + TransportRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            ).count()
            + VisaApplication.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            ).count()
            + AccommodationRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            ).count()
        )

        # Calculate percentage change
        total_change = (
            ((total_requests - previous_total) / previous_total * 100)
            if previous_total > 0
            else 0
        )

        # Calculate average processing time (from workflows)
        completed_workflows = WorkflowInstance.objects.filter(
            status="approved", completed_at__gte=start_date
        )

        avg_processing_hours = 0
        if completed_workflows.exists():
            processing_times = []
            for wf in completed_workflows:
                if wf.started_at and wf.completed_at:
                    delta = wf.completed_at - wf.started_at
                    processing_times.append(
                        delta.total_seconds() / 3600
                    )  # Convert to hours

            if processing_times:
                avg_processing_hours = sum(processing_times) / len(processing_times)

        # Previous period processing time
        previous_completed_workflows = WorkflowInstance.objects.filter(
            status="approved",
            completed_at__gte=previous_start,
            completed_at__lt=start_date,
        )

        previous_avg_hours = 0
        if previous_completed_workflows.exists():
            prev_times = []
            for wf in previous_completed_workflows:
                if wf.started_at and wf.completed_at:
                    delta = wf.completed_at - wf.started_at
                    prev_times.append(delta.total_seconds() / 3600)
            if prev_times:
                previous_avg_hours = sum(prev_times) / len(prev_times)

        processing_change = (
            ((avg_processing_hours - previous_avg_hours) / previous_avg_hours * 100)
            if previous_avg_hours > 0
            else 0
        )

        # Calculate completion rate
        total_workflows = WorkflowInstance.objects.filter(
            started_at__gte=start_date
        ).count()
        completed_count = completed_workflows.count()
        completion_rate = (
            (completed_count / total_workflows * 100) if total_workflows > 0 else 0
        )

        # Previous period completion rate
        prev_total_workflows = WorkflowInstance.objects.filter(
            started_at__gte=previous_start, started_at__lt=start_date
        ).count()
        prev_completed = WorkflowInstance.objects.filter(
            status="approved",
            completed_at__gte=previous_start,
            completed_at__lt=start_date,
        ).count()
        prev_completion_rate = (
            (prev_completed / prev_total_workflows * 100)
            if prev_total_workflows > 0
            else 0
        )
        completion_change = completion_rate - prev_completion_rate

        # Pending requests
        pending_count = (
            trf_requests.exclude(
                status__in=["Approved", "Rejected", "Cancelled"]
            ).count()
            + transport_requests.exclude(
                status__in=["Approved", "Rejected", "Cancelled"]
            ).count()
            + visa_requests.exclude(
                status__in=["Approved", "Rejected", "Cancelled"]
            ).count()
            + accommodation_requests.exclude(
                status__in=["Approved", "Rejected", "Cancelled"]
            ).count()
        )

        previous_pending = (
            TravelRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            )
            .exclude(status__in=["Approved", "Rejected", "Cancelled"])
            .count()
            + TransportRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            )
            .exclude(status__in=["Approved", "Rejected", "Cancelled"])
            .count()
            + VisaApplication.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            )
            .exclude(status__in=["Approved", "Rejected", "Cancelled"])
            .count()
            + AccommodationRequest.objects.filter(
                created_at__gte=previous_start, created_at__lt=start_date
            )
            .exclude(status__in=["Approved", "Rejected", "Cancelled"])
            .count()
        )

        pending_change = (
            ((pending_count - previous_pending) / previous_pending * 100)
            if previous_pending > 0
            else 0
        )

        # Requests by type
        requests_by_type = {
            "labels": ["Travel", "Transport", "Visa", "Accommodation"],
            "data": [
                trf_requests.count(),
                transport_requests.count(),
                visa_requests.count(),
                accommodation_requests.count(),
            ],
        }

        # Processing time by type
        processing_by_type = {
            "labels": ["Travel", "Transport", "Visa", "Accommodation"],
            "data": [],
        }

        # Calculate processing time for each type
        for model, entity_type in [
            (TravelRequest, "travelrequest"),
            (TransportRequest, "transportrequest"),
            (VisaApplication, "visaapplication"),
            (AccommodationRequest, "accommodation"),
        ]:
            workflows = WorkflowInstance.objects.filter(
                workflow_template__entity_type=entity_type,
                status="approved",
                completed_at__gte=start_date,
            )

            if workflows.exists():
                times = []
                for wf in workflows:
                    if wf.started_at and wf.completed_at:
                        delta = wf.completed_at - wf.started_at
                        times.append(delta.total_seconds() / 3600)
                avg_time = sum(times) / len(times) if times else 0
                processing_by_type["data"].append(round(avg_time, 1))
            else:
                processing_by_type["data"].append(0)

        # Request trends (last 12 months)
        months_data = self._get_monthly_trends()

        # Department statistics
        department_stats = self._get_department_stats(start_date)

        # Top performers (users with most approvals)
        top_performers = self._get_top_performers(start_date)

        report_data = {
            "key_metrics": [
                {
                    "name": "Total Requests",
                    "value": total_requests,
                    "change": round(total_change, 1),
                    "trend": (
                        "up"
                        if total_change > 0
                        else "down" if total_change < 0 else "neutral"
                    ),
                    "icon": "bi-file-earmark-text",
                },
                {
                    "name": "Avg. Processing Time",
                    "value": round(avg_processing_hours, 0),
                    "change": round(processing_change, 1),
                    "trend": (
                        "down"
                        if processing_change < 0
                        else "up" if processing_change > 0 else "neutral"
                    ),
                    "icon": "bi-clock",
                },
                {
                    "name": "Completion Rate",
                    "value": round(completion_rate, 0),
                    "change": round(completion_change, 1),
                    "trend": (
                        "up"
                        if completion_change > 0
                        else "down" if completion_change < 0 else "neutral"
                    ),
                    "icon": "bi-check-circle",
                },
                {
                    "name": "Pending Requests",
                    "value": pending_count,
                    "change": round(pending_change, 1),
                    "trend": (
                        "down"
                        if pending_change < 0
                        else "up" if pending_change > 0 else "neutral"
                    ),
                    "icon": "bi-hourglass-split",
                },
            ],
            "requests_by_type": requests_by_type,
            "processing_by_type": processing_by_type,
            "monthly_trends": months_data,
            "department_stats": department_stats,
            "top_performers": top_performers,
        }

        return success_response(
            data=report_data,
            message="Admin reports retrieved successfully",
            status_code=200,
        )

    def _get_monthly_trends(self):
        """Get request trends for last 12 months"""
        now = timezone.now()
        months = []
        submitted_data = []
        completed_data = []

        for i in range(11, -1, -1):
            month_start = timezone.datetime(
                now.year, now.month, 1, tzinfo=timezone.get_current_timezone()
            ) - timedelta(days=30 * i)
            month_end = month_start + timedelta(days=30)

            months.append(month_start.strftime("%b"))

            # Count submitted requests
            submitted = (
                TravelRequest.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).count()
                + TransportRequest.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).count()
                + VisaApplication.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).count()
                + AccommodationRequest.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).count()
            )
            submitted_data.append(submitted)

            # Count completed workflows
            completed = WorkflowInstance.objects.filter(
                status="approved",
                completed_at__gte=month_start,
                completed_at__lt=month_end,
            ).count()
            completed_data.append(completed)

        return {
            "labels": months,
            "submitted": submitted_data,
            "completed": completed_data,
        }

    def _get_department_stats(self, start_date):
        """Get statistics by department"""
        from accounts.models import Department

        # Get distinct department IDs from users
        department_ids = User.objects.values_list("department", flat=True).distinct()
        stats = []

        for dept_id in department_ids:
            if not dept_id:
                continue

            # Look up department name
            try:
                department = Department.objects.get(id=dept_id)
                dept_name = department.name
            except Department.DoesNotExist:
                dept_name = str(dept_id)  # Fallback to ID if not found

            # Get all requests from users in this department
            dept_users = User.objects.filter(department=dept_id)

            total = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users, created_at__gte=start_date
                ).count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users, created_at__gte=start_date
                ).count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date
                ).count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users, created_at__gte=start_date
                ).count()
            )

            if total == 0:
                continue

            pending = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
            )

            completed = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date, status="Approved"
                ).count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
            )

            # Calculate average processing time for this department
            dept_workflows = WorkflowInstance.objects.filter(
                initiated_by__in=dept_users,
                status="approved",
                completed_at__gte=start_date,
            )

            avg_time = 0
            if dept_workflows.exists():
                times = []
                for wf in dept_workflows:
                    if wf.started_at and wf.completed_at:
                        delta = wf.completed_at - wf.started_at
                        times.append(delta.total_seconds() / 3600)
                if times:
                    avg_time = sum(times) / len(times)

            stats.append(
                {
                    "department": dept_name,
                    "total": total,
                    "pending": pending,
                    "completed": completed,
                    "avgProcessingTime": round(avg_time, 1),
                }
            )

        # Sort by total requests
        stats.sort(key=lambda x: x["total"], reverse=True)
        return stats

    def _get_top_performers(self, start_date):
        """Get top performing approvers"""
        # Get users who have approved steps
        approvers = (
            WorkflowStepExecution.objects.filter(
                status="approved",
                action_date__gte=start_date,
                actioned_by__isnull=False,
            )
            .values("actioned_by")
            .annotate(processed=Count("id"))
            .order_by("-processed")[:5]
        )

        performers = []
        for approver in approvers:
            user = User.objects.get(id=approver["actioned_by"])

            # Calculate average processing time for this user
            user_steps = WorkflowStepExecution.objects.filter(
                actioned_by=user, status="approved", action_date__gte=start_date
            )

            times = []
            for step in user_steps:
                if step.created_at and step.action_date:
                    delta = step.action_date - step.created_at
                    times.append(delta.total_seconds() / 3600)

            avg_time = sum(times) / len(times) if times else 0

            performers.append(
                {
                    "name": user.get_full_name() or user.email,
                    "processed": approver["processed"],
                    "avgTime": round(avg_time, 1),
                }
            )

        return performers


class DepartmentalReportsView(APIView):
    """
    Detailed departmental reports
    Shows comprehensive statistics for each department
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get detailed department reports
        Query params:
        - department: specific department name (optional)
        - date_range: week, month, quarter, year (default: month)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        date_range = request.query_params.get("date_range", "month")
        department_filter = request.query_params.get("department")

        # Calculate date range
        now = timezone.now()
        if date_range == "week":
            start_date = now - timedelta(days=7)
        elif date_range == "quarter":
            start_date = now - timedelta(days=90)
        elif date_range == "year":
            start_date = now - timedelta(days=365)
        else:  # month
            start_date = now - timedelta(days=30)

        # Get departments
        from accounts.models import Department

        department_ids = User.objects.values_list("department", flat=True).distinct()
        if department_filter:
            department_ids = [department_filter]

        reports = []
        for dept_id in department_ids:
            if not dept_id:
                continue

            # Look up department name
            try:
                department = Department.objects.get(id=dept_id)
                dept_name = department.name
            except Department.DoesNotExist:
                dept_name = str(dept_id)  # Fallback to ID if not found

            dept_users = User.objects.filter(department=dept_id)

            # Get all request counts by type
            trf_count = TravelRequest.objects.filter(
                created_by__in=dept_users, created_at__gte=start_date
            ).count()

            transport_count = TransportRequest.objects.filter(
                requestor__in=dept_users, created_at__gte=start_date
            ).count()

            visa_count = VisaApplication.objects.filter(
                user__in=dept_users, created_at__gte=start_date
            ).count()

            accommodation_count = AccommodationRequest.objects.filter(
                trf__created_by__in=dept_users, created_at__gte=start_date
            ).count()

            total_requests = (
                trf_count + transport_count + visa_count + accommodation_count
            )

            # Get status breakdown
            approved = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date, status="Approved"
                ).count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Approved",
                ).count()
            )

            pending = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users, created_at__gte=start_date
                )
                .exclude(status__in=["Approved", "Rejected", "Cancelled"])
                .count()
            )

            rejected = (
                TravelRequest.objects.filter(
                    created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Rejected",
                ).count()
                + TransportRequest.objects.filter(
                    requestor__in=dept_users,
                    created_at__gte=start_date,
                    status="Rejected",
                ).count()
                + VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date, status="Rejected"
                ).count()
                + AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users,
                    created_at__gte=start_date,
                    status="Rejected",
                ).count()
            )

            # Calculate average processing time
            dept_workflows = WorkflowInstance.objects.filter(
                initiated_by__in=dept_users,
                status="approved",
                completed_at__gte=start_date,
            )

            avg_processing_time = 0
            if dept_workflows.exists():
                times = []
                for wf in dept_workflows:
                    if wf.started_at and wf.completed_at:
                        delta = wf.completed_at - wf.started_at
                        times.append(delta.total_seconds() / 3600)
                if times:
                    avg_processing_time = sum(times) / len(times)

            # Get top requestors in this department
            top_requestors = []
            for user in dept_users[:10]:  # Top 10 users
                user_requests = (
                    TravelRequest.objects.filter(
                        created_by=user, created_at__gte=start_date
                    ).count()
                    + TransportRequest.objects.filter(
                        requestor=user, created_at__gte=start_date
                    ).count()
                    + VisaApplication.objects.filter(
                        user=user, created_at__gte=start_date
                    ).count()
                )
                if user_requests > 0:
                    top_requestors.append(
                        {
                            "name": user.get_full_name() or user.email,
                            "email": user.email,
                            "requests": user_requests,
                        }
                    )

            top_requestors.sort(key=lambda x: x["requests"], reverse=True)
            top_requestors = top_requestors[:5]  # Keep top 5

            reports.append(
                {
                    "department": dept_name,
                    "totalRequests": total_requests,
                    "breakdown": {
                        "travel": trf_count,
                        "transport": transport_count,
                        "visa": visa_count,
                        "accommodation": accommodation_count,
                    },
                    "status": {
                        "approved": approved,
                        "pending": pending,
                        "rejected": rejected,
                    },
                    "avgProcessingTime": round(avg_processing_time, 1),
                    "topRequestors": top_requestors,
                    "activeUsers": dept_users.count(),
                }
            )

        # Sort by total requests
        reports.sort(key=lambda x: x["totalRequests"], reverse=True)

        return success_response(
            data={
                "reports": reports,
                "dateRange": date_range,
                "startDate": start_date.isoformat(),
                "endDate": now.isoformat(),
            },
            message="Departmental reports retrieved successfully",
            status_code=200,
        )


class UserActivityReportsView(APIView):
    """
    User activity reports
    Shows detailed activity for individual users or all users
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get user activity reports
        Query params:
        - user_id: specific user ID (optional)
        - date_range: week, month, quarter, year (default: month)
        - activity_type: requests, approvals, all (default: all)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        date_range = request.query_params.get("date_range", "month")
        user_id = request.query_params.get("user_id")
        activity_type = request.query_params.get("activity_type", "all")

        # Calculate date range
        now = timezone.now()
        if date_range == "week":
            start_date = now - timedelta(days=7)
        elif date_range == "quarter":
            start_date = now - timedelta(days=90)
        elif date_range == "year":
            start_date = now - timedelta(days=365)
        else:  # month
            start_date = now - timedelta(days=30)

        # Get users
        if user_id:
            try:
                users = [User.objects.get(id=user_id)]
            except User.DoesNotExist:
                return error_response(message="User not found", status_code=404)
        else:
            users = User.objects.all()[:50]  # Limit to 50 users for performance

        user_activities = []
        for user in users:
            # Get department name
            dept_name = user.department.name if user.department else None

            activity = {
                "userId": user.id,
                "name": user.get_full_name() or user.email,
                "email": user.email,
                "department": dept_name,
            }

            # Requests submitted
            if activity_type in ["requests", "all"]:
                travel_requests = TravelRequest.objects.filter(
                    created_by=user, created_at__gte=start_date
                ).count()

                transport_requests = TransportRequest.objects.filter(
                    requestor=user, created_at__gte=start_date
                ).count()

                visa_requests = VisaApplication.objects.filter(
                    user=user, created_at__gte=start_date
                ).count()

                activity["requestsSubmitted"] = {
                    "total": travel_requests + transport_requests + visa_requests,
                    "travel": travel_requests,
                    "transport": transport_requests,
                    "visa": visa_requests,
                }

            # Approvals processed
            if activity_type in ["approvals", "all"]:
                approvals = WorkflowStepExecution.objects.filter(
                    actioned_by=user, action_date__gte=start_date
                )

                approved_count = approvals.filter(status="approved").count()
                rejected_count = approvals.filter(status="rejected").count()
                pending_count = approvals.filter(status="pending").count()

                # Calculate average approval time
                approval_times = []
                for approval in approvals.filter(status__in=["approved", "rejected"]):
                    if approval.created_at and approval.action_date:
                        delta = approval.action_date - approval.created_at
                        approval_times.append(delta.total_seconds() / 3600)

                avg_approval_time = (
                    sum(approval_times) / len(approval_times) if approval_times else 0
                )

                activity["approvalsProcessed"] = {
                    "total": approvals.count(),
                    "approved": approved_count,
                    "rejected": rejected_count,
                    "pending": pending_count,
                    "avgTimeHours": round(avg_approval_time, 1),
                }

            # Last activity
            last_request = (
                TravelRequest.objects.filter(created_by=user)
                .order_by("-created_at")
                .first()
            )
            last_approval = (
                WorkflowStepExecution.objects.filter(actioned_by=user)
                .order_by("-action_date")
                .first()
            )

            last_activity_date = None
            if last_request and last_approval:
                last_activity_date = max(
                    last_request.created_at,
                    last_approval.action_date or last_approval.created_at,
                )
            elif last_request:
                last_activity_date = last_request.created_at
            elif last_approval:
                last_activity_date = (
                    last_approval.action_date or last_approval.created_at
                )

            activity["lastActivityDate"] = (
                last_activity_date.isoformat() if last_activity_date else None
            )

            user_activities.append(activity)

        # Sort by total activity (requests + approvals)
        user_activities.sort(
            key=lambda x: x.get("requestsSubmitted", {}).get("total", 0)
            + x.get("approvalsProcessed", {}).get("total", 0),
            reverse=True,
        )

        return success_response(
            data={
                "users": user_activities,
                "dateRange": date_range,
                "activityType": activity_type,
                "startDate": start_date.isoformat(),
                "endDate": now.isoformat(),
            },
            message="User activity reports retrieved successfully",
            status_code=200,
        )


class FinancialSummaryReportsView(APIView):
    """
    Financial summary reports
    Shows estimated costs and budget tracking
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get financial summary reports
        Query params:
        - date_range: week, month, quarter, year (default: month)
        - department: specific department (optional)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        date_range = request.query_params.get("date_range", "month")
        department_filter = request.query_params.get("department")

        # Calculate date range
        now = timezone.now()
        if date_range == "week":
            start_date = now - timedelta(days=7)
        elif date_range == "quarter":
            start_date = now - timedelta(days=90)
        elif date_range == "year":
            start_date = now - timedelta(days=365)
        else:  # month
            start_date = now - timedelta(days=30)

        if department_filter:
            dept_users = User.objects.filter(department=department_filter)
        else:
            dept_users = None

        # Flight bookings costs
        flight_query = FlightBooking.objects.filter(created_at__gte=start_date)
        if dept_users:
            # Get TRFs from department users
            dept_trfs = TravelRequest.objects.filter(created_by__in=dept_users)
            flight_query = flight_query.filter(trf__in=dept_trfs)

        flight_cost = flight_query.aggregate(total=Sum("cost"))["total"] or 0

        # Hotel bookings costs
        hotel_query = HotelBooking.objects.filter(created_at__gte=start_date)
        if dept_users:
            hotel_query = hotel_query.filter(trf__in=dept_trfs)

        hotel_cost = hotel_query.aggregate(total=Sum("cost_per_night"))["total"] or 0

        # Transport requests (estimated costs)
        transport_query = TransportRequest.objects.filter(created_at__gte=start_date)
        if dept_users:
            transport_query = transport_query.filter(requestor__in=dept_users)

        # Estimate transport cost based on count (placeholder - adjust as needed)
        transport_count = transport_query.count()
        estimated_transport_cost = (
            transport_count * 50
        )  # $50 per transport request estimate

        # Visa applications (estimated costs)
        visa_query = VisaApplication.objects.filter(created_at__gte=start_date)
        if dept_users:
            visa_query = visa_query.filter(user__in=dept_users)

        visa_count = visa_query.count()
        estimated_visa_cost = visa_count * 200  # $200 per visa estimate

        # Accommodation requests (estimated costs)
        accommodation_query = AccommodationRequest.objects.filter(
            created_at__gte=start_date
        )
        if dept_users:
            accommodation_query = accommodation_query.filter(
                trf__created_by__in=dept_users
            )

        # Calculate nights and cost
        accommodation_cost = 0
        for acc in accommodation_query:
            if acc.check_in_date and acc.check_out_date:
                nights = (acc.check_out_date - acc.check_in_date).days
                accommodation_cost += nights * 100  # $100 per night estimate

        # Total costs
        total_cost = (
            flight_cost
            + hotel_cost
            + estimated_transport_cost
            + estimated_visa_cost
            + accommodation_cost
        )

        # Get breakdown by department
        from accounts.models import Department

        department_breakdown = []
        department_ids = User.objects.values_list("department", flat=True).distinct()

        for dept_id in department_ids:
            if not dept_id:
                continue

            # Look up department name
            try:
                department = Department.objects.get(id=dept_id)
                dept_name = department.name
            except Department.DoesNotExist:
                dept_name = str(dept_id)  # Fallback to ID if not found

            dept_users = User.objects.filter(department=dept_id)
            dept_trfs = TravelRequest.objects.filter(created_by__in=dept_users)

            dept_flights = (
                FlightBooking.objects.filter(
                    trf__in=dept_trfs, created_at__gte=start_date
                ).aggregate(total=Sum("cost"))["total"]
                or 0
            )

            dept_hotels = (
                HotelBooking.objects.filter(
                    trf__in=dept_trfs, created_at__gte=start_date
                ).aggregate(total=Sum("cost_per_night"))["total"]
                or 0
            )

            dept_transport = (
                TransportRequest.objects.filter(
                    requestor__in=dept_users, created_at__gte=start_date
                ).count()
                * 50
            )

            dept_visa = (
                VisaApplication.objects.filter(
                    user__in=dept_users, created_at__gte=start_date
                ).count()
                * 200
            )

            dept_accommodation = (
                AccommodationRequest.objects.filter(
                    trf__created_by__in=dept_users, created_at__gte=start_date
                ).count()
                * 300
            )  # Simplified estimate

            dept_total = (
                dept_flights
                + dept_hotels
                + dept_transport
                + dept_visa
                + dept_accommodation
            )

            if dept_total > 0:
                department_breakdown.append(
                    {
                        "department": dept_name,
                        "totalCost": round(dept_total, 2),
                        "breakdown": {
                            "flights": round(dept_flights, 2),
                            "hotels": round(dept_hotels, 2),
                            "transport": round(dept_transport, 2),
                            "visa": round(dept_visa, 2),
                            "accommodation": round(dept_accommodation, 2),
                        },
                    }
                )

        department_breakdown.sort(key=lambda x: x["totalCost"], reverse=True)

        # Monthly trend
        monthly_costs = []
        for i in range(11, -1, -1):
            month_start = timezone.datetime(
                now.year, now.month, 1, tzinfo=timezone.get_current_timezone()
            ) - timedelta(days=30 * i)
            month_end = month_start + timedelta(days=30)

            month_flights = (
                FlightBooking.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).aggregate(total=Sum("cost"))["total"]
                or 0
            )

            month_hotels = (
                HotelBooking.objects.filter(
                    created_at__gte=month_start, created_at__lt=month_end
                ).aggregate(total=Sum("cost_per_night"))["total"]
                or 0
            )

            monthly_costs.append(
                {
                    "month": month_start.strftime("%b %Y"),
                    "cost": round(month_flights + month_hotels, 2),
                }
            )

        return success_response(
            data={
                "summary": {
                    "totalCost": round(total_cost, 2),
                    "breakdown": {
                        "flights": round(flight_cost, 2),
                        "hotels": round(hotel_cost, 2),
                        "transport": round(estimated_transport_cost, 2),
                        "visa": round(estimated_visa_cost, 2),
                        "accommodation": round(accommodation_cost, 2),
                    },
                    "currency": "USD",  # Placeholder
                },
                "departmentBreakdown": department_breakdown,
                "monthlyTrend": monthly_costs,
                "dateRange": date_range,
                "startDate": start_date.isoformat(),
                "endDate": now.isoformat(),
            },
            message="Financial summary reports retrieved successfully",
            status_code=200,
        )


class ReportExportView(APIView):
    """
    Export reports to CSV, Excel, or PDF
    """

    permission_classes = [IsAuthenticated]
    # Disable DRF's format suffix negotiation to prevent conflict with our 'format' parameter
    format_kwarg = None

    def get(self, request):
        """
        Export report data
        Query params:
        - report_type: admin, departmental, user_activity, financial (required)
        - export_format: csv, excel, pdf (default: csv) - Note: use 'export_format' not 'format' to avoid DRF conflict
        - date_range: week, month, quarter, year (default: month)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        report_type = request.query_params.get("report_type")
        # Use 'export_format' parameter name to avoid conflict with DRF's 'format' content negotiation
        export_format = request.query_params.get("export_format", "csv")

        if not report_type:
            return error_response(
                message="report_type parameter is required", status_code=400
            )

        # Get report data based on type
        if report_type == "admin":
            view = AdminReportsView()
        elif report_type == "departmental":
            view = DepartmentalReportsView()
        elif report_type == "user_activity":
            view = UserActivityReportsView()
        elif report_type == "financial":
            view = FinancialSummaryReportsView()
        else:
            return error_response(
                message=f"Invalid report_type: {report_type}", status_code=400
            )

        # Get the report data
        try:
            response = view.get(request)
        except Exception as e:
            return error_response(
                message=f"Error generating report: {str(e)}", status_code=500
            )

        data = response.data.get("data", {})

        # Generate export based on format
        if export_format == "excel":
            return self._export_excel(report_type, data)
        elif export_format == "pdf":
            return self._export_pdf(report_type, data)
        else:  # csv
            return self._export_csv(report_type, data)

    def _export_csv(self, report_type, data):
        """Export data to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Add headers and rows based on report type
        if report_type == "departmental":
            writer.writerow(
                [
                    "Department",
                    "Total Requests",
                    "Travel",
                    "Transport",
                    "Visa",
                    "Accommodation",
                    "Approved",
                    "Pending",
                    "Rejected",
                    "Avg Processing Time (hrs)",
                    "Active Users",
                ]
            )

            for report in data.get("reports", []):
                writer.writerow(
                    [
                        report["department"],
                        report["totalRequests"],
                        report["breakdown"]["travel"],
                        report["breakdown"]["transport"],
                        report["breakdown"]["visa"],
                        report["breakdown"]["accommodation"],
                        report["status"]["approved"],
                        report["status"]["pending"],
                        report["status"]["rejected"],
                        report["avgProcessingTime"],
                        report["activeUsers"],
                    ]
                )

        elif report_type == "user_activity":
            writer.writerow(
                [
                    "Name",
                    "Email",
                    "Department",
                    "Requests Submitted",
                    "Approvals Processed",
                    "Last Activity",
                ]
            )

            for user in data.get("users", []):
                writer.writerow(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        user.get("requestsSubmitted", {}).get("total", 0),
                        user.get("approvalsProcessed", {}).get("total", 0),
                        user.get("lastActivityDate", ""),
                    ]
                )

        elif report_type == "financial":
            writer.writerow(["Category", "Cost (USD)"])
            summary = data.get("summary", {})
            breakdown = summary.get("breakdown", {})

            writer.writerow(["Total", summary.get("totalCost", 0)])
            writer.writerow(["Flights", breakdown.get("flights", 0)])
            writer.writerow(["Hotels", breakdown.get("hotels", 0)])
            writer.writerow(["Transport", breakdown.get("transport", 0)])
            writer.writerow(["Visa", breakdown.get("visa", 0)])
            writer.writerow(["Accommodation", breakdown.get("accommodation", 0)])

            writer.writerow([])
            writer.writerow(["Department", "Total Cost"])
            for dept in data.get("departmentBreakdown", []):
                writer.writerow([dept["department"], dept["totalCost"]])

        else:  # admin report
            writer.writerow(["Metric", "Value"])
            for metric in data.get("key_metrics", []):
                writer.writerow([metric["name"], metric["value"]])

        # Create HTTP response
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.csv"'
        )
        return response

    def _export_excel(self, report_type, data):
        """Export data to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"

        # Style definitions
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers and rows based on report type
        if report_type == "departmental":
            headers = [
                "Department",
                "Total Requests",
                "Travel",
                "Transport",
                "Visa",
                "Accommodation",
                "Approved",
                "Pending",
                "Rejected",
                "Avg Processing Time (hrs)",
                "Active Users",
            ]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Add data rows
            for report in data.get("reports", []):
                ws.append(
                    [
                        report["department"],
                        report["totalRequests"],
                        report["breakdown"]["travel"],
                        report["breakdown"]["transport"],
                        report["breakdown"]["visa"],
                        report["breakdown"]["accommodation"],
                        report["status"]["approved"],
                        report["status"]["pending"],
                        report["status"]["rejected"],
                        report["avgProcessingTime"],
                        report["activeUsers"],
                    ]
                )

        elif report_type == "user_activity":
            headers = [
                "Name",
                "Email",
                "Department",
                "Requests Submitted",
                "Approvals Processed",
                "Last Activity",
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for user in data.get("users", []):
                ws.append(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        user.get("requestsSubmitted", {}).get("total", 0),
                        user.get("approvalsProcessed", {}).get("total", 0),
                        user.get("lastActivityDate", ""),
                    ]
                )

        elif report_type == "financial":
            headers = ["Category", "Cost (USD)"]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            summary = data.get("summary", {})
            breakdown = summary.get("breakdown", {})

            ws.append(["Total", summary.get("totalCost", 0)])
            ws.append(["Flights", breakdown.get("flights", 0)])
            ws.append(["Hotels", breakdown.get("hotels", 0)])
            ws.append(["Transport", breakdown.get("transport", 0)])
            ws.append(["Visa", breakdown.get("visa", 0)])
            ws.append(["Accommodation", breakdown.get("accommodation", 0)])

        else:  # admin report
            headers = ["Metric", "Value"]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for metric in data.get("key_metrics", []):
                ws.append([metric["name"], metric["value"]])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.xlsx"'
        )
        return response

    def _export_pdf(self, report_type, data):
        """Export data to PDF format"""
        output = io.BytesIO()

        # Use landscape for wider tables
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
            textColor=colors.HexColor("#0d9488"),
        )

        # Build PDF content
        elements = []

        # Add title
        title_text = f"TMS {report_type.replace('_', ' ').title()} Report"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 12))

        # Table style
        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9fafb")],
                ),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )

        # Build table based on report type
        if report_type == "departmental":
            headers = [
                "Department",
                "Total",
                "Travel",
                "Transport",
                "Visa",
                "Accom.",
                "Approved",
                "Pending",
                "Rejected",
                "Avg Time (hrs)",
                "Users",
            ]
            table_data = [headers]

            for report in data.get("reports", []):
                table_data.append(
                    [
                        report["department"],
                        str(report["totalRequests"]),
                        str(report["breakdown"]["travel"]),
                        str(report["breakdown"]["transport"]),
                        str(report["breakdown"]["visa"]),
                        str(report["breakdown"]["accommodation"]),
                        str(report["status"]["approved"]),
                        str(report["status"]["pending"]),
                        str(report["status"]["rejected"]),
                        str(report["avgProcessingTime"]),
                        str(report["activeUsers"]),
                    ]
                )

        elif report_type == "user_activity":
            headers = [
                "Name",
                "Email",
                "Department",
                "Requests",
                "Approvals",
                "Last Activity",
            ]
            table_data = [headers]

            for user in data.get("users", []):
                table_data.append(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        str(user.get("requestsSubmitted", {}).get("total", 0)),
                        str(user.get("approvalsProcessed", {}).get("total", 0)),
                        (
                            user.get("lastActivityDate", "")[:10]
                            if user.get("lastActivityDate")
                            else ""
                        ),
                    ]
                )

        elif report_type == "financial":
            headers = ["Category", "Cost (USD)"]
            table_data = [headers]

            summary = data.get("summary", {})
            breakdown = summary.get("breakdown", {})

            table_data.append(["Total", f"${summary.get('totalCost', 0):,.2f}"])
            table_data.append(["Flights", f"${breakdown.get('flights', 0):,.2f}"])
            table_data.append(["Hotels", f"${breakdown.get('hotels', 0):,.2f}"])
            table_data.append(["Transport", f"${breakdown.get('transport', 0):,.2f}"])
            table_data.append(["Visa", f"${breakdown.get('visa', 0):,.2f}"])
            table_data.append(
                ["Accommodation", f"${breakdown.get('accommodation', 0):,.2f}"]
            )

            # Add department breakdown if available
            if data.get("departmentBreakdown"):
                table_data.append(["", ""])  # Empty row
                table_data.append(["Department Breakdown", ""])
                for dept in data.get("departmentBreakdown", []):
                    table_data.append(
                        [dept["department"], f"${dept['totalCost']:,.2f}"]
                    )

        else:  # admin report
            headers = ["Metric", "Value", "Change", "Trend"]
            table_data = [headers]

            for metric in data.get("key_metrics", []):
                trend_symbol = (
                    "↑"
                    if metric["trend"] == "up"
                    else "↓" if metric["trend"] == "down" else "→"
                )
                table_data.append(
                    [
                        metric["name"],
                        str(metric["value"]),
                        f"{metric['change']}%",
                        trend_symbol,
                    ]
                )

            # Add requests by type
            if data.get("requests_by_type"):
                table_data.append(["", "", "", ""])
                table_data.append(["Requests by Type", "", "", ""])
                labels = data["requests_by_type"].get("labels", [])
                values = data["requests_by_type"].get("data", [])
                for i, label in enumerate(labels):
                    value = values[i] if i < len(values) else 0
                    table_data.append([label, str(value), "", ""])

            # Add department stats
            if data.get("department_stats"):
                table_data.append(["", "", "", ""])
                table_data.append(["Department Statistics", "", "", ""])
                for dept in data.get("department_stats", [])[:5]:  # Top 5
                    completion_rate = (
                        round((dept["completed"] / dept["total"] * 100))
                        if dept["total"] > 0
                        else 0
                    )
                    table_data.append(
                        [
                            dept["department"],
                            str(dept["total"]),
                            f"{completion_rate}%",
                            f"{dept['avgProcessingTime']}h",
                        ]
                    )

        # Create table
        if table_data:
            # Calculate column widths based on content
            col_count = len(table_data[0]) if table_data else 1
            available_width = 10 * inch  # Landscape width minus margins
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)
            table.setStyle(table_style)
            elements.append(table)

        # Add footer with timestamp
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            "Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey
        )
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(
            Paragraph(
                f"Generated on {timestamp} | Travel Management System", footer_style
            )
        )

        # Build PDF
        doc.build(elements)
        output.seek(0)

        # Create HTTP response
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.pdf"'
        )
        return response
