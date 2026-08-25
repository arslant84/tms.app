"""
ReportExportView - exports admin/departmental/user-activity report data
to CSV, Excel, or PDF.

Split out of reports/views.py (see docs/CODEBASE_REFACTOR_ROADMAP.md
item 3) - a pure file move, no logic changed. Delegates to the other
three report views (moved to their own sibling modules in the same
split) to fetch the underlying report data before exporting it.
"""

import csv
import io

from django.db.models import Avg
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from utils.api_response import error_response

from .admin_reports_view import AdminReportsView
from .departmental_reports_view import DepartmentalReportsView
from .permissions import _require_admin_reports_permission
from .user_activity_reports_view import UserActivityReportsView


class ReportExportView(APIView):
    """
    Export reports to CSV, Excel, or PDF
    """

    permission_classes = [IsAuthenticated]
    # Disable DRF's format suffix negotiation to prevent conflict with our 'format' parameter
    format_kwarg = None

    def get(self, request):
        """
        Export report data
        Query params:
        - report_type: admin, departmental, user_activity (required)
        - export_format: csv, excel, pdf (default: csv) - Note: use 'export_format' not 'format' to avoid DRF conflict
        - date_range: week, month, quarter, year (default: month)
        """
        forbidden = _require_admin_reports_permission(request)
        if forbidden:
            return forbidden

        report_type = request.query_params.get("report_type")
        # Use 'export_format' parameter name to avoid conflict with DRF's 'format' content negotiation
        export_format = request.query_params.get("export_format", "csv")

        if not report_type:
            return error_response(
                message="report_type parameter is required", status_code=400
            )

        # Get report data based on type
        if report_type == "admin":
            view = AdminReportsView()
        elif report_type == "departmental":
            view = DepartmentalReportsView()
        elif report_type == "user_activity":
            view = UserActivityReportsView()
        else:
            return error_response(
                message=f"Invalid report_type: {report_type}", status_code=400
            )

        # Get the report data
        try:
            response = view.get(request)
        except Exception as e:
            return error_response(
                message=f"Error generating report: {str(e)}", status_code=500
            )

        data = response.data.get("data", {})

        # Generate export based on format
        if export_format == "excel":
            return self._export_excel(report_type, data)
        elif export_format == "pdf":
            return self._export_pdf(report_type, data)
        else:  # csv
            return self._export_csv(report_type, data)

    def _export_csv(self, report_type, data):
        """Export data to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Add headers and rows based on report type
        if report_type == "departmental":
            writer.writerow(
                [
                    "Department",
                    "Total Requests",
                    "Travel",
                    "Transport",
                    "Visa",
                    "Accommodation",
                    "Approved",
                    "Pending",
                    "Rejected",
                    "Avg Processing Time (hrs)",
                    "Active Users",
                ]
            )

            for report in data.get("reports", []):
                writer.writerow(
                    [
                        report["department"],
                        report["totalRequests"],
                        report["breakdown"]["travel"],
                        report["breakdown"]["transport"],
                        report["breakdown"]["visa"],
                        report["breakdown"]["accommodation"],
                        report["status"]["approved"],
                        report["status"]["pending"],
                        report["status"]["rejected"],
                        report["avgProcessingTime"],
                        report["activeUsers"],
                    ]
                )

        elif report_type == "user_activity":
            writer.writerow(
                [
                    "Name",
                    "Email",
                    "Department",
                    "Requests Submitted",
                    "Approvals Processed",
                    "Last Activity",
                ]
            )

            for user in data.get("users", []):
                writer.writerow(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        user.get("requestsSubmitted", {}).get("total", 0),
                        user.get("approvalsProcessed", {}).get("total", 0),
                        user.get("lastActivityDate", ""),
                    ]
                )

        else:  # admin report
            writer.writerow(["Metric", "Value"])
            for metric in data.get("key_metrics", []):
                writer.writerow([metric["name"], metric["value"]])

        # Create HTTP response
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.csv"'
        )
        return response

    def _export_excel(self, report_type, data):
        """Export data to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"

        # Style definitions
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers and rows based on report type
        if report_type == "departmental":
            headers = [
                "Department",
                "Total Requests",
                "Travel",
                "Transport",
                "Visa",
                "Accommodation",
                "Approved",
                "Pending",
                "Rejected",
                "Avg Processing Time (hrs)",
                "Active Users",
            ]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Add data rows
            for report in data.get("reports", []):
                ws.append(
                    [
                        report["department"],
                        report["totalRequests"],
                        report["breakdown"]["travel"],
                        report["breakdown"]["transport"],
                        report["breakdown"]["visa"],
                        report["breakdown"]["accommodation"],
                        report["status"]["approved"],
                        report["status"]["pending"],
                        report["status"]["rejected"],
                        report["avgProcessingTime"],
                        report["activeUsers"],
                    ]
                )

        elif report_type == "user_activity":
            headers = [
                "Name",
                "Email",
                "Department",
                "Requests Submitted",
                "Approvals Processed",
                "Last Activity",
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for user in data.get("users", []):
                ws.append(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        user.get("requestsSubmitted", {}).get("total", 0),
                        user.get("approvalsProcessed", {}).get("total", 0),
                        user.get("lastActivityDate", ""),
                    ]
                )

        else:  # admin report
            headers = ["Metric", "Value"]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for metric in data.get("key_metrics", []):
                ws.append([metric["name"], metric["value"]])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.xlsx"'
        )
        return response

    def _export_pdf(self, report_type, data):
        """Export data to PDF format"""
        output = io.BytesIO()

        # Use landscape for wider tables
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
            textColor=colors.HexColor("#0d9488"),
        )

        # Build PDF content
        elements = []

        # Add title
        title_text = f"TMS {report_type.replace('_', ' ').title()} Report"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 12))

        # Table style
        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9fafb")],
                ),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )

        # Build table based on report type
        if report_type == "departmental":
            headers = [
                "Department",
                "Total",
                "Travel",
                "Transport",
                "Visa",
                "Accom.",
                "Approved",
                "Pending",
                "Rejected",
                "Avg Time (hrs)",
                "Users",
            ]
            table_data = [headers]

            for report in data.get("reports", []):
                table_data.append(
                    [
                        report["department"],
                        str(report["totalRequests"]),
                        str(report["breakdown"]["travel"]),
                        str(report["breakdown"]["transport"]),
                        str(report["breakdown"]["visa"]),
                        str(report["breakdown"]["accommodation"]),
                        str(report["status"]["approved"]),
                        str(report["status"]["pending"]),
                        str(report["status"]["rejected"]),
                        str(report["avgProcessingTime"]),
                        str(report["activeUsers"]),
                    ]
                )

        elif report_type == "user_activity":
            headers = [
                "Name",
                "Email",
                "Department",
                "Requests",
                "Approvals",
                "Last Activity",
            ]
            table_data = [headers]

            for user in data.get("users", []):
                table_data.append(
                    [
                        user["name"],
                        user["email"],
                        user.get("department", ""),
                        str(user.get("requestsSubmitted", {}).get("total", 0)),
                        str(user.get("approvalsProcessed", {}).get("total", 0)),
                        (
                            user.get("lastActivityDate", "")[:10]
                            if user.get("lastActivityDate")
                            else ""
                        ),
                    ]
                )

        else:  # admin report
            headers = ["Metric", "Value", "Change", "Trend"]
            table_data = [headers]

            for metric in data.get("key_metrics", []):
                trend_symbol = (
                    "↑"
                    if metric["trend"] == "up"
                    else "↓" if metric["trend"] == "down" else "→"
                )
                table_data.append(
                    [
                        metric["name"],
                        str(metric["value"]),
                        f"{metric['change']}%",
                        trend_symbol,
                    ]
                )

            # Add requests by type
            if data.get("requests_by_type"):
                table_data.append(["", "", "", ""])
                table_data.append(["Requests by Type", "", "", ""])
                labels = data["requests_by_type"].get("labels", [])
                values = data["requests_by_type"].get("data", [])
                for i, label in enumerate(labels):
                    value = values[i] if i < len(values) else 0
                    table_data.append([label, str(value), "", ""])

            # Add department stats
            if data.get("department_stats"):
                table_data.append(["", "", "", ""])
                table_data.append(["Department Statistics", "", "", ""])
                for dept in data.get("department_stats", [])[:5]:  # Top 5
                    completion_rate = (
                        round((dept["completed"] / dept["total"] * 100))
                        if dept["total"] > 0
                        else 0
                    )
                    table_data.append(
                        [
                            dept["department"],
                            str(dept["total"]),
                            f"{completion_rate}%",
                            f"{dept['avgProcessingTime']}h",
                        ]
                    )

        # Create table
        if table_data:
            # Calculate column widths based on content
            col_count = len(table_data[0]) if table_data else 1
            available_width = 10 * inch  # Landscape width minus margins
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)
            table.setStyle(table_style)
            elements.append(table)

        # Add footer with timestamp
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            "Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey
        )
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(
            Paragraph(
                f"Generated on {timestamp} | Travel Management System", footer_style
            )
        )

        # Build PDF
        doc.build(elements)
        output.seek(0)

        # Create HTTP response
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{report_type}_report.pdf"'
        )
        return response
